# encoding=utf-8
'''
Filename : tool_box.py
Datatime : 2023/01/25 
Author :KJH-x
'''
import re
import sys
import os
import json
from json import JSONDecodeError
import zlib

menu_msg = "\
[1]: Refresh reference data and calculate\n\
[2]: Exit\n\
Choose :"

mirror_url = "https://pypi.tuna.tsinghua.edu.cn/simple"

check_time = 15

while 1:
    try:
        import pyperclip
        import openpyxl
        break
    except ImportError as ex:
        if check_time:
            print("try installing essential module now.")
            module_name = re.sub(
                "\\'", "", str(
                    re.findall("\\'.*?\\'$", str(ex))[0]
                ))
            os.system(
                f"{sys.executable} -m pip install {module_name} -i {mirror_url}"
            )
            check_time -= 1
        else:
            print("Allow retry times exceeded, exiting in 3s...")
            import time
            time.sleep(3)
            exit()


del check_time, mirror_url
os.chdir(sys.path[0])

data = {
    '@type': '@penguin-statistics/depot',
    'items': [
        {'id': 'id', 'have': 0, 'name': 'name'},
    ]
}
datalist = [
    [
        # item name, value amount, value per item, value grand total
        '物品名称', '拥有数量', '单件价值', '总价值',
    ]

]
reference = [
    {
        'itemId': 'id',
        'itemName': 'name',
        'itemType': 'quality',
        'itemValueGreen': 0,
        'itemValueReason': 0,
        'version': 'version',
    }
]
ref_list = [
    {'name': '_test', 'value': 0}
]


def convert_ref():
    for i in reference:
        ref_list.append(
            {
                'name': i['itemName'],
                'value': float(i['itemValueGreen'])
            }
        )
    return


def convert_src():
    for i in data['items']:
        datalist.append(
            [
                i['name'],
                str(i['have'])
            ]
        )
    return


def calculate_value() -> float:
    value = 0
    for check in datalist:
        for ref in ref_list:
            if check[0] == ref['name']:
                item_value = int(check[1])*ref['value']
                check.append(ref['value'])
                check.append(item_value)
                value += item_value

    return value


def replenish():
    for check in datalist:
        if len(check) != 4:
            for dif in range(4-len(check)):
                check.append('0')


def write2xlsx(data: list[list[str]], CRC32: str) -> None:
    data_book = openpyxl.Workbook()
    try:
        data_book = openpyxl.open(".\\detail.xlsx", read_only=False)
    except FileNotFoundError:
        data_book = openpyxl.Workbook()
    data_sheet_1 = data_book.create_sheet(f'材料价值统计{CRC32}')
    for rowY in range(len(data)):
        for colX in range(4):
            data_sheet_1.cell(
                row=rowY+1, column=colX+1,
                value=str(data[rowY][colX])
            )
    while 1:
        try:
            data_book.save(".\\detail.xlsx")
            break
        except PermissionError:
            input("The table sheet might be opened by another program.\nClose it and type enter to retry data write")
    return


def textCRC(context: str) -> str:
    return f"{(zlib.crc32(context.encode('utf8')) & 0xffffffff):08X}"


try:
    while 1:
        select = re.match("[0-9]{1}", input(menu_msg))
        if select:
            timestamp = ""
            match int(select.group()):
                case 1:
                    os.popen("python .\\refresh_reference.py")
                    with open(".\\material_value.json", 'r', encoding='utf8')as ref_file:
                        reference = json.load(ref_file)
                    try:
                        import pyperclip
                        clipboard = pyperclip.paste()
                        data = json.loads(clipboard)
                        timestamp = textCRC(clipboard)

                    except JSONDecodeError:
                        print("Cannot decode clipboard")
                        while 1:
                            try:
                                storage_json = input("Select a json file.")
                                text = ""
                                with open(storage_json, 'r', encoding='utf8') as json_content:
                                    data = json.load(json_content)
                                    text = json_content.read()
                                timestamp = textCRC(text)
                                break
                            except FileNotFoundError:
                                pass
                    convert_ref()
                    convert_src()
                    value_GT = calculate_value()
                    print(
                        f'The Total value of data is {value_GT:.3f}\n',
                        f'Equivalent to {value_GT/135:.3f} stone.'
                    )
                    replenish()
                    write2xlsx(datalist, timestamp)
                    print()
                case 2:
                    exit()  
                case _: 
                    print("error")
        else:
            print("Not support syntax, check input.")
            pass
except KeyboardInterrupt:
    exit()
